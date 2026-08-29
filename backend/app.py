from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import sqlite3
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import requests

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*", "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"]}}, supports_credentials=True)

# Database setup
DATABASE = '/app/data/wealth_manager.db'

def get_db():
    """Get database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# API configuration for API Ninjas (stock price and currency conversion)
API_NINJAS_KEY = os.environ.get('API_NINJAS_KEY') or ''
API_NINJAS_HEADERS = {'X-Api-Key': API_NINJAS_KEY} if API_NINJAS_KEY else {}

def get_current_stock_price(symbol):
    """Fetch current stock price (USD) for a given ticker symbol."""
    try:
        if not API_NINJAS_KEY:
            # No API key configured: skip live price fetch and use a safe fallback
            return 0.0

        url = f'https://api.api-ninjas.com/v1/stockprice?ticker={symbol}'
        resp = requests.get(url, headers=API_NINJAS_HEADERS, timeout=10)
        if resp.status_code == requests.codes.ok:
            data = resp.json()
            return float(data.get('price', 0.0))
        else:
            print(f"Error fetching price for {symbol}: {resp.status_code}")
            return 0.0
    except Exception as e:
        print(f"Exception fetching price for {symbol}: {e}")
        return 0.0

def convert_usd_to_inr(amount):
    """Convert USD amount to INR using API Ninjas convertcurrency endpoint."""
    try:
        if not API_NINJAS_KEY:
            # Rough fallback conversion rate when no API key is configured
            return round(amount * 83.0, 2)

        convert_url = f'https://api.api-ninjas.com/v1/convertcurrency?have=USD&want=INR&amount={amount}'
        resp = requests.get(convert_url, headers=API_NINJAS_HEADERS, timeout=10)
        if resp.status_code == requests.codes.ok:
            data = resp.json()
            return float(data.get('new_amount', 0.0))
        else:
            print(f"Error converting USD to INR: {resp.status_code}")
            return round(amount * 83.0, 2)
    except Exception as e:
        print(f"Exception converting currency: {e}")
        return round(amount * 83.0, 2)

def init_db():
    """Initialize database with tables"""
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Fixed Deposits table - Create if not exists
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS fixed_deposits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_name TEXT NOT NULL,
            cust_id TEXT,
            fd_number TEXT,
            principal REAL NOT NULL,
            maturity_amt REAL,
            interest_amt REAL,
            rate REAL NOT NULL,
            tenure_years INTEGER NOT NULL,
            maturity_date TEXT NOT NULL,
            date_created TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Add missing columns if they don't exist (migration support)
    try:
        cursor.execute('ALTER TABLE fixed_deposits ADD COLUMN cust_id TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE fixed_deposits ADD COLUMN fd_number TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE fixed_deposits ADD COLUMN maturity_amt REAL')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE fixed_deposits ADD COLUMN interest_amt REAL')
    except:
        pass
    
    # Mutual Funds table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mutual_funds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fund_name TEXT NOT NULL,
            units REAL NOT NULL,
            nav REAL NOT NULL,
            total_value REAL NOT NULL,
            purchase_date TEXT NOT NULL,
            date_created TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Stocks table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stocks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stock_name TEXT NOT NULL,
            symbol TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            total_value REAL NOT NULL,
            purchase_date TEXT NOT NULL,
            date_created TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Migration: Remove purchase_price or current_price columns if they exist (SQLite doesn't support DROP COLUMN directly)
    try:
        # Check if purchase_price column exists
        cursor.execute("PRAGMA table_info(stocks)")
        columns = [row[1] for row in cursor.fetchall()]
        if 'purchase_price' in columns or 'current_price' in columns:
            # Create new table without purchase_price/current_price
            cursor.execute('''
                CREATE TABLE stocks_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    stock_name TEXT NOT NULL,
                    symbol TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    total_value REAL NOT NULL,
                    purchase_date TEXT NOT NULL,
                    date_created TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            # Copy data; ignore current_price if present
            cursor.execute('''
                INSERT INTO stocks_new (id, stock_name, symbol, quantity, total_value, purchase_date, date_created)
                SELECT id, stock_name, symbol, quantity, total_value, purchase_date, date_created FROM stocks
            ''')
            # Drop old table and rename new one
            cursor.execute('DROP TABLE stocks')
            cursor.execute('ALTER TABLE stocks_new RENAME TO stocks')
    except Exception as e:
        print(f"Migration warning for stocks table: {e}")
    
    # RBI Bonds table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rbi_bonds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bond_type TEXT NOT NULL,
            bond_number TEXT NOT NULL,
            amount REAL NOT NULL,
            rate REAL NOT NULL,
            tenure_years INTEGER NOT NULL,
            maturity_date TEXT NOT NULL,
            purchase_date TEXT NOT NULL,
            date_created TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Add missing columns if they don't exist (migration support for rbi_bonds)
    try:
        cursor.execute('ALTER TABLE rbi_bonds ADD COLUMN bond_number TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE rbi_bonds ADD COLUMN tenure_years INTEGER')
    except:
        pass
    
    # PPF (Public Provident Fund) table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ppf (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_number TEXT NOT NULL,
            financial_year TEXT NOT NULL,
            amount REAL NOT NULL,
            rate REAL NOT NULL,
            maturity_year INTEGER NOT NULL,
            date_of_investment TEXT NOT NULL,
            date_created TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Add date_of_investment column if it doesn't exist (for existing databases)
    try:
        cursor.execute('ALTER TABLE ppf ADD COLUMN date_of_investment TEXT NOT NULL DEFAULT \'\'')
    except:
        pass  # Column already exists
    
    # Portfolio Snapshots table - for tracking historical data
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS portfolio_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_date TEXT NOT NULL,
            fixed_deposits_total REAL NOT NULL,
            mutual_funds_total REAL NOT NULL,
            stocks_total REAL NOT NULL,
            rbi_bonds_total REAL NOT NULL,
            ppf_total REAL NOT NULL,
            total_portfolio_value REAL NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

# Routes for Fixed Deposits
@app.route('/api/fixed-deposits', methods=['GET'])
def get_fixed_deposits():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM fixed_deposits ORDER BY maturity_date ASC')
    deposits = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(deposits)

@app.route('/api/fixed-deposits', methods=['POST'])
def add_fixed_deposit():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO fixed_deposits (bank_name, cust_id, fd_number, principal, maturity_amt, interest_amt, rate, tenure_years, maturity_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (data['bank_name'], data['cust_id'], data['fd_number'], data['principal'], data['maturity_amt'], data['interest_amt'], data['rate'], data['tenure_years'], data['maturity_date']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Fixed deposit added successfully'}), 201

@app.route('/api/fixed-deposits/<int:fd_id>', methods=['PUT'])
def update_fixed_deposit(fd_id):
    try:
        data = request.json
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM fixed_deposits WHERE id = ?', (fd_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Fixed deposit not found'}), 404
        
        cursor.execute('''
            UPDATE fixed_deposits 
            SET bank_name=?, cust_id=?, fd_number=?, principal=?, maturity_amt=?, interest_amt=?, rate=?, tenure_years=?, maturity_date=?
            WHERE id=?
        ''', (data['bank_name'], data['cust_id'], data['fd_number'], data['principal'], data['maturity_amt'], data['interest_amt'], data['rate'], data['tenure_years'], data['maturity_date'], fd_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Fixed deposit updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/fixed-deposits/<int:fd_id>', methods=['DELETE'])
def delete_fixed_deposit(fd_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM fixed_deposits WHERE id = ?', (fd_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Fixed deposit not found'}), 404
        
        cursor.execute('DELETE FROM fixed_deposits WHERE id = ?', (fd_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Fixed deposit deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Routes for Mutual Funds
@app.route('/api/mutual-funds', methods=['GET'])
def get_mutual_funds():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM mutual_funds')
    funds = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(funds)

@app.route('/api/mutual-funds', methods=['POST'])
def add_mutual_fund():
    data = request.json
    total_value = data['units'] * data['nav']
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO mutual_funds (fund_name, units, nav, total_value, purchase_date)
        VALUES (?, ?, ?, ?, ?)
    ''', (data['fund_name'], data['units'], data['nav'], total_value, data['purchase_date']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Mutual fund added successfully'}), 201

@app.route('/api/mutual-funds/<int:mf_id>', methods=['PUT'])
def update_mutual_fund(mf_id):
    try:
        data = request.json
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM mutual_funds WHERE id = ?', (mf_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Mutual fund not found'}), 404
        
        total_value = data['units'] * data['nav']
        cursor.execute('''
            UPDATE mutual_funds 
            SET fund_name=?, units=?, nav=?, total_value=?, purchase_date=?
            WHERE id=?
        ''', (data['fund_name'], data['units'], data['nav'], total_value, data['purchase_date'], mf_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Mutual fund updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mutual-funds/<int:mf_id>', methods=['DELETE'])
def delete_mutual_fund(mf_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM mutual_funds WHERE id = ?', (mf_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Mutual fund not found'}), 404
        
        cursor.execute('DELETE FROM mutual_funds WHERE id = ?', (mf_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Mutual fund deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Routes for Stocks
def enrich_stock_with_current_values(stock):
    """Return stock record with current_price and total_value computed from API."""
    symbol = stock['symbol']
    quantity = stock['quantity']
    total_value_stored = stock['total_value']

    price_usd = get_current_stock_price(symbol)
    if price_usd > 0 and quantity > 0:
        total_usd = round(price_usd * quantity, 2)
        total_inr = convert_usd_to_inr(total_usd)
        current_price = round(total_inr / quantity, 2)
    else:
        total_inr = total_value_stored
        current_price = round(total_inr / quantity, 2) if quantity else 0.0
        total_usd = 0.0

    return {
        'id': stock['id'],
        'stock_name': stock['stock_name'],
        'symbol': symbol,
        'quantity': quantity,
        'current_price': current_price,
        'total_value': total_inr,
        'price_usd': price_usd,
        'total_usd': total_usd
    }

@app.route('/api/stocks', methods=['GET'])
def get_stocks():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM stocks')
    stocks = [enrich_stock_with_current_values(dict(row)) for row in cursor.fetchall()]
    conn.close()
    return jsonify(stocks)

@app.route('/api/stocks', methods=['POST'])
def add_stock():
    data = request.json
    symbol = data['symbol']
    quantity = data['quantity']
    price_usd = get_current_stock_price(symbol)
    total_value = 0.0
    if price_usd > 0 and quantity > 0:
        total_usd = round(price_usd * quantity, 2)
        total_value = convert_usd_to_inr(total_usd)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO stocks (stock_name, symbol, quantity, total_value, purchase_date)
        VALUES (?, ?, ?, ?, ?)
    ''', (data['stock_name'], symbol, quantity, total_value, data['purchase_date']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Stock added successfully'}), 201

@app.route('/api/stocks/<int:stock_id>', methods=['PUT'])
def update_stock(stock_id):
    try:
        data = request.json
        symbol = data['symbol']
        quantity = data['quantity']
        price_usd = get_current_stock_price(symbol)
        total_value = 0.0
        if price_usd > 0 and quantity > 0:
            total_usd = round(price_usd * quantity, 2)
            total_value = convert_usd_to_inr(total_usd)

        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM stocks WHERE id = ?', (stock_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Stock not found'}), 404
        
        cursor.execute('''
            UPDATE stocks 
            SET stock_name=?, symbol=?, quantity=?, total_value=?, purchase_date=?
            WHERE id=?
        ''', (data['stock_name'], symbol, quantity, total_value, data['purchase_date'], stock_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Stock updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stocks/<int:stock_id>', methods=['DELETE'])
def delete_stock(stock_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM stocks WHERE id = ?', (stock_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Stock not found'}), 404
        
        cursor.execute('DELETE FROM stocks WHERE id = ?', (stock_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Stock deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stocks/current-values', methods=['GET'])
def get_stocks_current_values():
    """Refresh current stock total values from API Ninjas and update stored totals."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, stock_name, symbol, quantity, total_value FROM stocks')
        stocks = [dict(row) for row in cursor.fetchall()]

        results = []
        for s in stocks:
            symbol = s['symbol']
            quantity = s['quantity']
            total_value_stored = s['total_value']
            price_usd = get_current_stock_price(symbol)
            total_usd = round(price_usd * quantity, 2) if price_usd > 0 and quantity > 0 else 0.0
            if price_usd > 0 and quantity > 0:
                total_inr = convert_usd_to_inr(total_usd)
                current_price = round(total_inr / quantity, 2)
                cursor.execute('UPDATE stocks SET total_value = ? WHERE id = ?', (total_inr, s['id']))
            else:
                total_inr = total_value_stored
                current_price = round(total_inr / quantity, 2) if quantity else 0.0

            results.append({
                'id': s['id'],
                'stock_name': s['stock_name'],
                'symbol': symbol,
                'quantity': quantity,
                'price_usd': price_usd,
                'current_price': current_price,
                'total_usd': total_usd,
                'total_value': total_inr
            })

        conn.commit()
        conn.close()
        return jsonify(results), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Routes for RBI Bonds
@app.route('/api/rbi-bonds', methods=['GET'])
def get_rbi_bonds():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM rbi_bonds ORDER BY purchase_date ASC')
    bonds = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(bonds)

@app.route('/api/rbi-bonds', methods=['POST'])
def add_rbi_bond():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO rbi_bonds (bond_type, bond_number, amount, rate, tenure_years, maturity_date, purchase_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (data['bond_type'], data['bond_number'], data['amount'], data['rate'], data['tenure_years'], 
          data['maturity_date'], data['purchase_date']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'RBI bond added successfully'}), 201

@app.route('/api/rbi-bonds/<int:rbi_bonds_id>', methods=['PUT'])
def update_rbi_bond(rbi_bonds_id):
    try:
        data = request.json
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM rbi_bonds WHERE id = ?', (rbi_bonds_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'RBI bond not found'}), 404
        
        cursor.execute('''
            UPDATE rbi_bonds 
            SET bond_type=?, bond_number=?, amount=?, rate=?, tenure_years=?, maturity_date=?, purchase_date=?
            WHERE id=?
        ''', (data['bond_type'], data['bond_number'], data['amount'], data['rate'], data['tenure_years'], data['maturity_date'], data['purchase_date'], rbi_bonds_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'RBI bond updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/rbi-bonds/<int:rbi_bonds_id>', methods=['DELETE'])
def delete_rbi_bond(rbi_bonds_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM rbi_bonds WHERE id = ?', (rbi_bonds_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'RBI bond not found'}), 404
        
        cursor.execute('DELETE FROM rbi_bonds WHERE id = ?', (rbi_bonds_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'RBI bond deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Routes for PPF
@app.route('/api/ppf', methods=['GET'])
def get_ppf():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM ppf ORDER BY date_of_investment ASC')
    ppf_records = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(ppf_records)

@app.route('/api/ppf', methods=['POST'])
def add_ppf():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO ppf (account_number, financial_year, amount, rate, maturity_year, date_of_investment)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (data['account_number'], data['financial_year'], data['amount'], data['rate'], data['maturity_year'], data['date_of_investment']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'PPF record added successfully'}), 201

@app.route('/api/ppf/<int:ppf_id>', methods=['PUT'])
def update_ppf(ppf_id):
    try:
        data = request.json
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM ppf WHERE id = ?', (ppf_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'PPF record not found'}), 404
        
        cursor.execute('''
            UPDATE ppf 
            SET account_number=?, financial_year=?, amount=?, rate=?, maturity_year=?, date_of_investment=?
            WHERE id=?
        ''', (data['account_number'], data['financial_year'], data['amount'], data['rate'], data['maturity_year'], data['date_of_investment'], ppf_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'PPF record updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ppf/<int:ppf_id>', methods=['DELETE'])
def delete_ppf(ppf_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM ppf WHERE id = ?', (ppf_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'PPF record not found'}), 404
        
        cursor.execute('DELETE FROM ppf WHERE id = ?', (ppf_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'PPF record deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Helper function to calculate total PPF interest earned
def get_ppf_total_interest():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT amount, rate, date_of_investment, financial_year FROM ppf WHERE amount > 0 AND rate > 0')
    ppf_records = cursor.fetchall()
    total_ppf_interest = 0
    
    from datetime import datetime
    current_year = datetime.now().year
    
    for record in ppf_records:
        amount = record['amount']
        rate = record['rate']
        date_of_investment = record['date_of_investment']
        financial_year = record['financial_year']
        
        total_ppf_interest = (total_ppf_interest + amount) * ((1 + (rate / 400)) ** (1 * 4))

    conn.close()
    return total_ppf_interest

# Portfolio Summary
@app.route('/api/portfolio-summary', methods=['GET'])
def get_portfolio_summary():
    conn = get_db()
    cursor = conn.cursor()
    
    summary = {}
    
    # Fixed Deposits Total
    cursor.execute('SELECT SUM(principal) as total FROM fixed_deposits')
    summary['fixed_deposits'] = cursor.fetchone()['total'] or 0
    
    # Mutual Funds Total
    cursor.execute('SELECT SUM(total_value) as total FROM mutual_funds')
    summary['mutual_funds'] = cursor.fetchone()['total'] or 0
    
    # Stocks Total
    cursor.execute('SELECT SUM(total_value) as total FROM stocks')
    summary['stocks'] = cursor.fetchone()['total'] or 0
    
    # RBI Bonds Total
    cursor.execute('SELECT SUM(amount) as total FROM rbi_bonds')
    summary['rbi_bonds'] = cursor.fetchone()['total'] or 0
    
    # PPF Total Interest Earned (calculate individually for accuracy)   
    summary['ppf'] = get_ppf_total_interest()
    
    summary['total_portfolio_value'] = sum(summary.values())
    #summary['total_portfolio_value'] = total_ppf_interest

    conn.close()
    return jsonify(summary)

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'Backend is running', 'timestamp': datetime.now().isoformat()})

# Portfolio Snapshots
@app.route('/api/portfolio-snapshot', methods=['POST'])
def record_portfolio_snapshot():
    """Record a snapshot of the current portfolio totals"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get current totals
        cursor.execute('SELECT SUM(principal) as total FROM fixed_deposits')
        fd_total = cursor.fetchone()['total'] or 0
        
        cursor.execute('SELECT SUM(total_value) as total FROM mutual_funds')
        mf_total = cursor.fetchone()['total'] or 0
        
        cursor.execute('SELECT SUM(total_value) as total FROM stocks')
        stocks_total = cursor.fetchone()['total'] or 0
        
        cursor.execute('SELECT SUM(amount) as total FROM rbi_bonds')
        bonds_total = cursor.fetchone()['total'] or 0
        
        #cursor.execute('SELECT SUM(amount) as total FROM ppf')
        ppf_total = get_ppf_total_interest()
        
        total_portfolio = fd_total + mf_total + stocks_total + bonds_total + ppf_total
        
        # Get snapshot date from request or use today
        data = request.json or {}
        snapshot_date = data.get('snapshot_date', datetime.now().strftime('%Y-%m-%d'))
        
        # Insert snapshot
        cursor.execute('''
            INSERT INTO portfolio_snapshots 
            (snapshot_date, fixed_deposits_total, mutual_funds_total, stocks_total, rbi_bonds_total, ppf_total, total_portfolio_value)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (snapshot_date, fd_total, mf_total, stocks_total, bonds_total, ppf_total, total_portfolio))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'Snapshot recorded successfully',
            'snapshot_date': snapshot_date,
            'total_portfolio_value': total_portfolio
        }), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/portfolio-snapshots', methods=['GET'])
def get_portfolio_snapshots():
    """Get all portfolio snapshots ordered by date"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM portfolio_snapshots 
            ORDER BY snapshot_date ASC
        ''')
        snapshots = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify(snapshots), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/portfolio-snapshots/<int:portfolio_id>', methods=['DELETE'])
def delete_portfolio_snapshot(portfolio_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if record exists
        cursor.execute('SELECT id FROM portfolio_snapshots WHERE id = ?', (portfolio_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'Portfolio snapshot not found'}), 404
        
        cursor.execute('DELETE FROM portfolio_snapshots WHERE id = ?', (portfolio_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Portfolio snapshot deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Excel Export Endpoint
def create_styled_header(ws, row, headers):
    """Create a styled header row in the worksheet"""
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

def add_data_rows(ws, start_row, data, headers):
    """Add data rows to worksheet with formatting"""
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, record in enumerate(data, start_row):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = record.get(header, '')
            cell.border = border
            
            # Center align and format currency columns
            if 'amount' in header.lower() or 'value' in header.lower() or 'price' in header.lower() or 'principal' in header.lower() or 'nav' in header.lower():
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

def adjust_column_widths(ws):
    """Adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export all wealth data to Excel with multiple sheets"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    create_styled_header(ws_summary, 1, ["Asset Type", "Total Value (₹)"])
    
    cursor.execute('SELECT SUM(principal) as total FROM fixed_deposits')
    fd_total = cursor.fetchone()['total'] or 0
    
    cursor.execute('SELECT SUM(total_value) as total FROM mutual_funds')
    mf_total = cursor.fetchone()['total'] or 0
    
    cursor.execute('SELECT SUM(total_value) as total FROM stocks')
    stocks_total = cursor.fetchone()['total'] or 0
    
    cursor.execute('SELECT SUM(amount) as total FROM rbi_bonds')
    bonds_total = cursor.fetchone()['total'] or 0
    
    # PPF Total Interest Earned (calculate individually for accuracy)
    cursor.execute('SELECT amount, rate, date_of_investment, financial_year FROM ppf WHERE amount > 0 AND rate > 0')
    ppf_records = cursor.fetchall()
    ppf_total = 0
    
    for record in ppf_records:
        amount = record['amount']
        rate = record['rate']
        date_of_investment = record['date_of_investment']
        financial_year = record['financial_year']
        
        ppf_total = (ppf_total + amount) * ((1 + (rate / 400)) ** (1 * 4))
    
    total_portfolio = fd_total + mf_total + stocks_total + bonds_total + ppf_total
    
    summary_data = [
        {"Asset Type": "Fixed Deposits", "Total Value (₹)": fd_total},
        {"Asset Type": "Mutual Funds", "Total Value (₹)": mf_total},
        {"Asset Type": "Stocks", "Total Value (₹)": stocks_total},
        {"Asset Type": "RBI Bonds", "Total Value (₹)": bonds_total},
        {"Asset Type": "PPF", "Total Value (₹)": ppf_total},
        {"Asset Type": "Total Portfolio", "Total Value (₹)": total_portfolio},
    ]
    
    add_data_rows(ws_summary, 2, summary_data, ["Asset Type", "Total Value (₹)"])
    adjust_column_widths(ws_summary)
    
    # Fixed Deposits Sheet
    cursor.execute('SELECT * FROM fixed_deposits')
    fd_data = [dict(row) for row in cursor.fetchall()]
    if fd_data:
        ws_fd = wb.create_sheet("Fixed Deposits")
        headers = list(fd_data[0].keys())
        create_styled_header(ws_fd, 1, headers)
        add_data_rows(ws_fd, 2, fd_data, headers)
        adjust_column_widths(ws_fd)
    
    # Mutual Funds Sheet
    cursor.execute('SELECT * FROM mutual_funds')
    mf_data = [dict(row) for row in cursor.fetchall()]
    if mf_data:
        ws_mf = wb.create_sheet("Mutual Funds")
        headers = list(mf_data[0].keys())
        create_styled_header(ws_mf, 1, headers)
        add_data_rows(ws_mf, 2, mf_data, headers)
        adjust_column_widths(ws_mf)
    
    # Stocks Sheet
    cursor.execute('SELECT * FROM stocks')
    stocks_data = [dict(row) for row in cursor.fetchall()]
    if stocks_data:
        ws_stocks = wb.create_sheet("Stocks")
        headers = list(stocks_data[0].keys())
        create_styled_header(ws_stocks, 1, headers)
        add_data_rows(ws_stocks, 2, stocks_data, headers)
        adjust_column_widths(ws_stocks)
    
    # RBI Bonds Sheet
    cursor.execute('SELECT * FROM rbi_bonds')
    bonds_data = [dict(row) for row in cursor.fetchall()]
    if bonds_data:
        ws_bonds = wb.create_sheet("RBI Bonds")
        headers = list(bonds_data[0].keys())
        create_styled_header(ws_bonds, 1, headers)
        add_data_rows(ws_bonds, 2, bonds_data, headers)
        adjust_column_widths(ws_bonds)
    
    # PPF Sheet
    cursor.execute('SELECT * FROM ppf')
    ppf_data = [dict(row) for row in cursor.fetchall()]
    if ppf_data:
        ws_ppf = wb.create_sheet("PPF")
        headers = list(ppf_data[0].keys())
        create_styled_header(ws_ppf, 1, headers)
        add_data_rows(ws_ppf, 2, ppf_data, headers)
        adjust_column_widths(ws_ppf)
    
    # Historical Data Sheet
    cursor.execute('SELECT * FROM portfolio_snapshots ORDER BY snapshot_date ASC')
    snapshots_data = [dict(row) for row in cursor.fetchall()]
    if snapshots_data:
        ws_history = wb.create_sheet("Historical Data")
        headers = ["Snapshot Date", "Fixed Deposits", "Mutual Funds", "Stocks", "RBI Bonds", "PPF", "Total Portfolio Value"]
        create_styled_header(ws_history, 1, headers)
        
        for row_idx, snapshot in enumerate(snapshots_data, 2):
            ws_history.cell(row=row_idx, column=1).value = snapshot['snapshot_date']
            ws_history.cell(row=row_idx, column=2).value = snapshot['fixed_deposits_total']
            ws_history.cell(row=row_idx, column=3).value = snapshot['mutual_funds_total']
            ws_history.cell(row=row_idx, column=4).value = snapshot['stocks_total']
            ws_history.cell(row=row_idx, column=5).value = snapshot['rbi_bonds_total']
            ws_history.cell(row=row_idx, column=6).value = snapshot['ppf_total']
            ws_history.cell(row=row_idx, column=7).value = snapshot['total_portfolio_value']
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in range(1, 8):
                cell = ws_history.cell(row=row_idx, column=col)
                cell.border = border
                if col > 1:  # Currency columns
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        adjust_column_widths(ws_history)
    
    conn.close()
    
    # Save to BytesIO and send
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Wealth_Manager_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    """Import data from an Excel file (same format as export) and replace current data."""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        conn = get_db()
        cursor = conn.cursor()

        def safe_float(v):
            try:
                return float(v) if v is not None and v != '' else 0
            except:
                return 0

        def safe_int(v):
            try:
                return int(v) if v is not None and v != '' else 0
            except:
                return 0

        # Fixed Deposits
        if 'Fixed Deposits' in wb.sheetnames:
            ws = wb['Fixed Deposits']
            rows = list(ws.values)
            if len(rows) > 1:
                headers = [str(h).strip() for h in rows[0]]
                cursor.execute('DELETE FROM fixed_deposits')
                for row in rows[1:]:
                    row_dict = dict(zip(headers, row))
                    cursor.execute('''
                        INSERT INTO fixed_deposits (bank_name, cust_id, fd_number, principal, maturity_amt, interest_amt, rate, tenure_years, maturity_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        row_dict.get('bank_name'),
                        row_dict.get('cust_id'),
                        row_dict.get('fd_number'),
                        safe_float(row_dict.get('principal')),
                        safe_float(row_dict.get('maturity_amt')),
                        safe_float(row_dict.get('interest_amt')),
                        safe_float(row_dict.get('rate')),
                        safe_int(row_dict.get('tenure_years')),
                        row_dict.get('maturity_date') or ''
                    ))

        # Mutual Funds
        if 'Mutual Funds' in wb.sheetnames:
            ws = wb['Mutual Funds']
            rows = list(ws.values)
            if len(rows) > 1:
                headers = [str(h).strip() for h in rows[0]]
                cursor.execute('DELETE FROM mutual_funds')
                for row in rows[1:]:
                    row_dict = dict(zip(headers, row))
                    cursor.execute('''
                        INSERT INTO mutual_funds (fund_name, units, nav, total_value, purchase_date)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        row_dict.get('fund_name'),
                        safe_float(row_dict.get('units')),
                        safe_float(row_dict.get('nav')),
                        safe_float(row_dict.get('total_value')),
                        row_dict.get('purchase_date') or ''
                    ))

        # Stocks
        if 'Stocks' in wb.sheetnames:
            ws = wb['Stocks']
            rows = list(ws.values)
            if len(rows) > 1:
                headers = [str(h).strip() for h in rows[0]]
                cursor.execute('DELETE FROM stocks')
                for row in rows[1:]:
                    row_dict = dict(zip(headers, row))
                    quantity_value = safe_int(row_dict.get('quantity'))
                    # Prefer imported total_value, else compute from current_price if available
                    total_value = safe_float(row_dict.get('total_value'))
                    if total_value == 0:
                        total_value = quantity_value * safe_float(row_dict.get('current_price'))
                    cursor.execute('''
                        INSERT INTO stocks (stock_name, symbol, quantity, total_value, purchase_date)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        row_dict.get('stock_name'),
                        row_dict.get('symbol'),
                        quantity_value,
                        total_value,
                        row_dict.get('purchase_date') or ''
                    ))

        # RBI Bonds
        if 'RBI Bonds' in wb.sheetnames:
            ws = wb['RBI Bonds']
            rows = list(ws.values)
            if len(rows) > 1:
                headers = [str(h).strip() for h in rows[0]]
                cursor.execute('DELETE FROM rbi_bonds')
                for row in rows[1:]:
                    row_dict = dict(zip(headers, row))
                    bond_type = row_dict.get('bond_type', '')
                    bond_number = row_dict.get('bond_number', '')
                    amount = safe_float(row_dict.get('amount'))
                    rate = safe_float(row_dict.get('rate'))
                    tenure_years = safe_int(row_dict.get('tenure_years'))
                    maturity_date = row_dict.get('maturity_date', '')
                    purchase_date = row_dict.get('purchase_date', '')
                    cursor.execute('''
                        INSERT INTO rbi_bonds (bond_type, bond_number, amount, rate, tenure_years, maturity_date, purchase_date)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (bond_type, bond_number, amount, rate, tenure_years, maturity_date, purchase_date))

        # PPF
        if 'PPF' in wb.sheetnames:
            ws = wb['PPF']
            rows = list(ws.values)
            if len(rows) > 1:
                headers = [str(h).strip() for h in rows[0]]
                cursor.execute('DELETE FROM ppf')
                for row in rows[1:]:
                    row_dict = dict(zip(headers, row))
                    account_number = row_dict.get('account_number') or row_dict.get('Account Number') or ''
                    financial_year = row_dict.get('financial_year') or row_dict.get('Financial Year') or ''
                    amount = safe_float(row_dict.get('amount'))
                    rate = safe_float(row_dict.get('rate'))
                    maturity_year = safe_int(row_dict.get('maturity_year'))
                    date_of_investment = row_dict.get('date_of_investment') or row_dict.get('Date of Investment') or ''
                    date_created = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute('''
                        INSERT INTO ppf (account_number, financial_year, amount, rate, maturity_year, date_of_investment, date_created)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (account_number, financial_year, amount, rate, maturity_year, date_of_investment, date_created))

        # Historical Data (portfolio_snapshots)
        if 'Historical Data' in wb.sheetnames:
            ws = wb['Historical Data']
            rows = list(ws.values)
            if len(rows) > 1:
                # headers expected: Snapshot Date, Fixed Deposits, Mutual Funds, Stocks, RBI Bonds, PPF, Total Portfolio Value
                cursor.execute('DELETE FROM portfolio_snapshots')
                for row in rows[1:]:
                    # row[0] snapshot_date, row[1] fd, row[2] mf, row[3] stocks, row[4] bonds, row[5] ppf, row[6] total
                    snapshot_date = None
                    if row[0] is not None:
                        if hasattr(row[0], 'strftime'):
                            snapshot_date = row[0].strftime('%Y-%m-%d')
                        else:
                            snapshot_date = str(row[0])
                    else:
                        snapshot_date = datetime.now().strftime('%Y-%m-%d')

                    fd_total = safe_float(row[1])
                    mf_total = safe_float(row[2])
                    stocks_total = safe_float(row[3])
                    bonds_total = safe_float(row[4])
                    ppf_total = safe_float(row[5])
                    total_portfolio = safe_float(row[6])

                    cursor.execute('''
                        INSERT INTO portfolio_snapshots (snapshot_date, fixed_deposits_total, mutual_funds_total, stocks_total, rbi_bonds_total, ppf_total, total_portfolio_value)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (snapshot_date, fd_total, mf_total, stocks_total, bonds_total, ppf_total, total_portfolio))

        conn.commit()
        conn.close()

        return jsonify({'message': 'Import successful'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
